#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Document-task input model.

This module intentionally lives next to, rather than inside, document_model.py.
The audit pipeline keeps its DOCX/PPTX-only behavior, while document tasks can
work over a broader corpus that includes XLSX workbooks and read-only PDF pages.
"""

from __future__ import annotations

from copy import deepcopy
from pathlib import Path
from typing import Any, Dict, Iterable, List

from document_model import (
    ProjectPaths,
    block_paragraph_refs_docx,
    build_block_map as build_ooxml_block_map,
    default_project_paths,
    ensure_project_dirs,
    rel_parent,
    source_relative,
)


SUPPORTED_TASK_EXTS = {".docx", ".pptx", ".xlsx", ".pdf"}


def is_supported_task_input(path: Path) -> bool:
    return path.suffix.lower() in SUPPORTED_TASK_EXTS and not path.name.startswith("~$")


def iter_task_inputs(input_dir: Path, *, recursive: bool = True) -> List[Path]:
    if input_dir.is_file():
        return [input_dir] if is_supported_task_input(input_dir) else []
    pattern = "**/*" if recursive else "*"
    return sorted(path for path in input_dir.glob(pattern) if path.is_file() and is_supported_task_input(path))


def safe_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", " ").replace("\n", " ").strip()
    return " ".join(text.split())


def short_cell_value(value: Any, *, max_len: int = 500) -> str:
    text = safe_text(value)
    if len(text) <= max_len:
        return text
    return text[: max_len - 1] + "..."


def column_letter(index: int) -> str:
    from openpyxl.utils import get_column_letter

    return get_column_letter(index)


def build_task_block_map(source: Path, input_dir: Path, *, pdf_max_pages: int = 5) -> Dict[str, Any]:
    ext = source.suffix.lower()
    if ext in {".docx", ".pptx"}:
        payload = build_ooxml_block_map(source, input_dir)
        payload["task_document_type"] = ext.lstrip(".")
        return payload
    if ext == ".xlsx":
        return build_xlsx_block_map(source, input_dir)
    if ext == ".pdf":
        return build_pdf_block_map(source, input_dir, max_pages=pdf_max_pages)
    raise ValueError(f"Unsupported document-task input: {source}")


def clean_pdf_text(text: str) -> str:
    lines = [" ".join(line.strip().split()) for line in str(text or "").replace("\r", "\n").split("\n")]
    lines = [line for line in lines if line]
    return "\n".join(lines)


def build_pdf_block_map(source: Path, input_dir: Path, *, max_pages: int = 5) -> Dict[str, Any]:
    try:
        import fitz
    except ImportError as exc:
        raise RuntimeError("PyMuPDF is required for PDF document-task input.") from exc

    page_limit = int(max_pages or 0)
    blocks: List[Dict[str, Any]] = []
    document = fitz.open(str(source))
    try:
        total_pages = len(document)
        if page_limit > 0:
            total_to_read = min(total_pages, page_limit)
        else:
            total_to_read = total_pages

        for page_index in range(total_to_read):
            page = document.load_page(page_index)
            text = clean_pdf_text(page.get_text("text"))
            if not text:
                continue
            page_number = page_index + 1
            blocks.append(
                {
                    "block_id": f"pdf_p_{page_number:04d}",
                    "marker": "",
                    "kind": "pdf_page",
                    "object_type": "pdf_page",
                    "text": text,
                    "technical_location": {
                        "part": "pdf",
                        "page": page_number,
                        "page_index": page_index,
                        "total_pages": total_pages,
                        "extracted_pages": total_to_read,
                    },
                    "location": f"Страница {page_number}",
                    "page": page_number,
                    "corpus_row_index": page_number,
                }
            )
    finally:
        document.close()

    return {
        "source_relative_path": source_relative(source, input_dir).as_posix(),
        "source_path": str(source),
        "document_type": "pdf",
        "task_document_type": "pdf",
        "pdf_max_pages": page_limit,
        "blocks": blocks,
    }


def build_xlsx_block_map(source: Path, input_dir: Path) -> Dict[str, Any]:
    from openpyxl import load_workbook

    workbook = load_workbook(str(source), data_only=False, read_only=True)
    blocks: List[Dict[str, Any]] = []

    for sheet_index, sheet in enumerate(workbook.worksheets, start=1):
        header_values: dict[int, str] = {}
        header_row_index = 0
        block_row_index = 0

        for row_index, row in enumerate(sheet.iter_rows(), start=1):
            cells: list[tuple[int, str]] = []
            for col_index, cell in enumerate(row, start=1):
                text = short_cell_value(cell.value)
                if text:
                    cells.append((col_index, text))
            if not cells:
                continue

            if not header_values:
                header_row_index = row_index
                header_values = {col_index: text for col_index, text in cells if text}

            block_row_index += 1
            parts: list[str] = []
            for col_index, text in cells:
                letter = column_letter(col_index)
                header = header_values.get(col_index, "")
                if header and row_index != header_row_index and header != text:
                    parts.append(f"{letter} ({header}) = {text}")
                else:
                    parts.append(f"{letter} = {text}")

            block_id = f"xlsx_s{sheet_index:03d}_r{row_index:06d}"
            blocks.append(
                {
                    "block_id": block_id,
                    "marker": "",
                    "kind": "worksheet_row",
                    "object_type": "worksheet_row",
                    "text": " | ".join(parts),
                    "technical_location": {
                        "part": "workbook",
                        "sheet": sheet.title,
                        "sheet_index": sheet_index,
                        "row_index": row_index,
                        "header_row_index": header_row_index,
                    },
                    "location": f"{sheet.title}!{row_index}:{row_index}",
                    "worksheet": sheet.title,
                    "row_index": row_index,
                    "corpus_row_index": block_row_index,
                }
            )

    workbook.close()
    return {
        "source_relative_path": source_relative(source, input_dir).as_posix(),
        "source_path": str(source),
        "document_type": "xlsx",
        "task_document_type": "xlsx",
        "blocks": blocks,
    }


def build_corpus_block_map(block_maps: Iterable[Dict[str, Any]]) -> Dict[str, Any]:
    documents: list[dict[str, Any]] = []
    blocks: list[dict[str, Any]] = []

    for doc_index, block_map in enumerate(block_maps, start=1):
        source_rel = str(block_map.get("source_relative_path") or "")
        document_type = str(block_map.get("document_type") or block_map.get("task_document_type") or "")
        prefix = f"d{doc_index:03d}"
        documents.append(
            {
                "document_index": doc_index,
                "source_relative_path": source_rel,
                "document_type": document_type,
                "blocks": len(block_map.get("blocks", []) or []),
            }
        )
        for block in block_map.get("blocks", []) or []:
            if not isinstance(block, dict):
                continue
            item = deepcopy(block)
            original_block_id = str(item.get("block_id") or "")
            item["block_id"] = f"{prefix}__{original_block_id}"
            item["original_block_id"] = original_block_id
            item["source_relative_path"] = source_rel
            item["source_document_type"] = document_type
            item["corpus_document_index"] = doc_index
            blocks.append(item)

    return {
        "source_relative_path": "__corpus__",
        "source_path": "",
        "document_type": "corpus",
        "task_document_type": "corpus",
        "documents": documents,
        "blocks": blocks,
    }


def block_lookup(block_map: Dict[str, Any]) -> dict[str, dict[str, Any]]:
    return {str(block.get("block_id") or ""): block for block in block_map.get("blocks", []) or [] if isinstance(block, dict)}
